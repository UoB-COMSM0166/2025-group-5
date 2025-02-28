from openpyxl import load_workbook
import json

jsonData = [];

workbook = load_workbook('levels.xlsx')
sheet = workbook['level1']
for row in range(1, 26):
    for col in range(1, 41):
        cell = sheet.cell(row=row, column=col)
        if cell.fill.patternType is None or cell.fill.patternType == 'none':
            continue
        else:
            new_data = dict()
            new_data['type'] = 'wall'
            new_data['position'] = dict()
            new_data['position']['x'] = 32 * col - 32
            new_data['position']['y'] = 32 * row - 32
            jsonData.append(new_data)

with open('output.json', 'w') as json_file:
    json.dump(jsonData, json_file, indent=4)
    json_file.close()